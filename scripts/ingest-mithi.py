"""Ingest Mithi River source files into public/asset (same filenames the UI expects)."""

from __future__ import annotations

import json
import math
import re
import shutil
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[1]
ASSET = ROOT / "public" / "asset"
DOWNLOADS = Path(r"C:\Users\Kunal.Desale\Downloads")
INGEST = ROOT / "_ingest"

ASSET.mkdir(parents=True, exist_ok=True)


def parse_pairs(text: str):
    out = []
    for pair in text.strip().split():
        parts = pair.split(",")
        if len(parts) < 2:
            continue
        lon, lat = float(parts[0]), float(parts[1])
        out.append([lon, lat])
    return out


def simple_data(block: str, key: str):
    m = re.search(rf'<SimpleData name="{key}">([^<]*)</SimpleData>', block)
    return m.group(1).strip() if m else None


def kml_box(kml: str):
    return {
        "north": float(re.search(r"<north>([^<]+)</north>", kml).group(1)),
        "south": float(re.search(r"<south>([^<]+)</south>", kml).group(1)),
        "east": float(re.search(r"<east>([^<]+)</east>", kml).group(1)),
        "west": float(re.search(r"<west>([^<]+)</west>", kml).group(1)),
    }


def coords_from_box(box):
    return [
        [box["west"], box["north"]],
        [box["east"], box["north"]],
        [box["east"], box["south"]],
        [box["west"], box["south"]],
    ]


def sample_overlay(png_path: Path, class_defs, max_d=48):
    im = Image.open(png_path).convert("RGBA")
    w, h = im.size
    opaque = [(r, g, b) for r, g, b, a in im.getdata() if a > 0]
    total = len(opaque)
    targets = []
    for d in class_defs:
        key = d["key"]
        targets.append((d["id"], int(key[0:2], 16), int(key[2:4], 16), int(key[4:6], 16)))
    buckets = Counter()
    max_d2 = max_d * max_d
    for r, g, b in opaque:
        best_id, best_d2 = None, max_d2 + 1
        for cid, tr, tg, tb in targets:
            d2 = (r - tr) ** 2 + (g - tg) ** 2 + (b - tb) ** 2
            if d2 < best_d2:
                best_d2 = d2
                best_id = cid
        if best_id is not None and best_d2 <= max_d2:
            buckets[best_id] += 1
    # Approximate ha from LatLonBox later; shares only here.
    classes = []
    for d in class_defs:
        px = buckets.get(d["id"], 0)
        classes.append(
            {
                **d,
                "class": d["id"],
                "pixels": px,
                "share_pct": round(100.0 * px / total, 1) if total else 0.0,
            }
        )
    return {
        "width": w,
        "height": h,
        "total_pixels": total,
        "classes": classes,
    }


def box_area_ha(box):
    # rough geographic rectangle area
    lat_m = (box["north"] - box["south"]) * 111_320
    mid = (box["north"] + box["south"]) / 2
    lon_m = (box["east"] - box["west"]) * 111_320 * math.cos(math.radians(mid))
    return (lat_m * lon_m) / 10_000


def convert_river():
    src = DOWNLOADS / "3875ff6842ca495e90921823311b1b6f.kml"
    dest = ASSET / "mula-mutha-river.kml"
    dest.write_bytes(src.read_bytes())
    # also keep a clear name
    (ASSET / "mithi-river.kml").write_bytes(src.read_bytes())
    print("river", dest.name)


def convert_buffer():
    src = DOWNLOADS / "2afcb94dc61646be8efe635167718230.kml"
    text = src.read_text(encoding="utf-8", errors="replace")
    features = []
    for m in re.finditer(r"<Placemark[\s\S]*?</Placemark>", text):
        block = m.group(0)
        coords_m = re.search(r"<coordinates>([\s\S]*?)</coordinates>", block)
        if not coords_m:
            continue
        ring = parse_pairs(coords_m.group(1))
        if len(ring) < 3:
            continue
        if ring[0] != ring[-1]:
            ring.append(ring[0])
        features.append(
            {
                "type": "Feature",
                "properties": {"name": "Mithi buffer", "id": simple_data(block, "id")},
                "geometry": {"type": "Polygon", "coordinates": [ring]},
            }
        )
    geo = {"type": "FeatureCollection", "features": features, "name": "Mithi buffered AOI"}
    (ASSET / "mithi-buffer.geojson").write_text(json.dumps(geo), encoding="utf-8")
    src_copy = ASSET / "mithi-buffer.kml"
    src_copy.write_bytes(src.read_bytes())
    print("buffer features", len(features))


def convert_garbage():
    src = DOWNLOADS / "34bf00ee982243b69bbdd2bd2e2432ca.kml"
    text = src.read_text(encoding="utf-8", errors="replace")
    features = []
    for m in re.finditer(r"<Placemark[\s\S]*?</Placemark>", text):
        block = m.group(0)
        name_m = re.search(r"<name>([^<]*)</name>", block)
        coords_m = re.search(r"<coordinates>([^<]+)</coordinates>", block)
        if not coords_m:
            continue
        lon, lat = map(float, coords_m.group(1).strip().split(",")[:2])
        features.append(
            {
                "type": "Feature",
                "properties": {
                    "id": int(name_m.group(1)) if name_m and name_m.group(1).isdigit() else len(features),
                    "name": name_m.group(1) if name_m else str(len(features)),
                    "kind": "garbage",
                },
                "geometry": {"type": "Point", "coordinates": [lon, lat]},
            }
        )
    geo = {
        "type": "FeatureCollection",
        "name": "Mithi detected garbage locations",
        "features": features,
    }
    (ASSET / "mula-mutha-garbage-locations.geojson").write_text(json.dumps(geo), encoding="utf-8")
    (ASSET / "mula-mutha-garbage-locations.kml").write_bytes(src.read_bytes())
    meta = {
        "name": "Mithi detected garbage locations",
        "count": len(features),
        "kind": "Estimated",
        "source_kml": src.name,
        "geojson": "/asset/mula-mutha-garbage-locations.geojson",
    }
    (ASSET / "mula-mutha-garbage-locations.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
    print("garbage", len(features))


def convert_tributaries():
    src = DOWNLOADS / "b08d34d8135b43b7bac4ceeabdee7d38.kml"
    text = src.read_text(encoding="utf-8", errors="replace")
    features = []
    class_counts = Counter()
    for m in re.finditer(r"<Placemark[\s\S]*?</Placemark>", text):
        block = m.group(0)
        name_m = re.search(r"<name>([^<]*)</name>", block)
        name = name_m.group(1).strip() if name_m else ""
        waterway = simple_data(block, "waterway") or ""
        osm_id = simple_data(block, "osm_id")
        coords_m = re.search(r"<coordinates>([\s\S]*?)</coordinates>", block)
        if not coords_m:
            continue
        line = parse_pairs(coords_m.group(1))
        if len(line) < 2:
            continue
        # classify like ClimateEye
        low = (name or "").lower()
        ww = waterway.lower()
        if ww in ("river",) or "mithi" in low:
            kind = "mainstem"
        elif ww in ("stream", "brook") or (name and name not in ("Untitled Path", "Mithi River")):
            kind = "stream"
        elif ww in ("drain", "ditch") or "nullah" in low or "nalla" in low:
            kind = "drain"
        elif ww == "canal":
            kind = "canal"
        elif ww == "ditch":
            kind = "ditch"
        elif not name or name == "Untitled Path":
            kind = "feeder"
        else:
            kind = "stream"
        class_counts[kind] += 1
        features.append(
            {
                "type": "Feature",
                "properties": {
                    "name": name or None,
                    "waterway": waterway or None,
                    "osm_id": osm_id,
                    "kind": kind,
                    "class": kind,
                },
                "geometry": {"type": "LineString", "coordinates": line},
            }
        )
    geo = {
        "type": "FeatureCollection",
        "name": "Mithi drainage / joining waterways",
        "features": features,
    }
    (ASSET / "mula-mutha-tributaries.geojson").write_text(json.dumps(geo), encoding="utf-8")
    (ASSET / "mula-mutha-tributaries.kml").write_bytes(src.read_bytes())
    print("tributaries", len(features), dict(class_counts))
    return class_counts


def convert_flood_water():
    src = DOWNLOADS / "611be66f88f4459da17e5e15ba498154.xlsx"
    wb = load_workbook(src, read_only=True, data_only=True)
    periods = []
    for i, row in enumerate(wb["areas"].iter_rows(values_only=True)):
        if i == 0:
            continue
        periods.append(
            {
                "id": i - 1,
                "pre_date": str(row[0]),
                "post_date": str(row[1]),
                "water_area_ha": round(float(row[2]), 2),
                "flood_area_ha": round(float(row[3]), 2),
            }
        )
    key_to_id = {(p["pre_date"], p["post_date"]): p["id"] for p in periods}
    grouped = {p["id"]: {"flood": [], "water": []} for p in periods}
    for i, row in enumerate(wb["lat_lon"].iter_rows(values_only=True)):
        if i == 0:
            continue
        pre, post, lat, lon, cls = row
        pid = key_to_id[(str(pre), str(post))]
        grouped[pid][cls].append([round(float(lon), 6), round(float(lat), 6)])
    wb.close()

    lons, lats = [], []
    # remove old period files beyond current count
    for old in ASSET.glob("mula-mutha-flood-water-*.json"):
        old.unlink()

    for pid, classes in grouped.items():
        periods[pid]["n_flood"] = len(classes["flood"])
        periods[pid]["n_water"] = len(classes["water"])
        periods[pid]["points"] = f"/asset/mula-mutha-flood-water-{pid}.json"
        for pair in classes["flood"] + classes["water"]:
            lons.append(pair[0])
            lats.append(pair[1])
        out = ASSET / f"mula-mutha-flood-water-{pid}.json"
        out.write_text(json.dumps(classes, separators=(",", ":")), encoding="utf-8")
        print("flood period", pid, "flood", periods[pid]["n_flood"], "water", periods[pid]["n_water"])

    peak = max(periods, key=lambda p: p["flood_area_ha"])
    index = {
        "name": "Mithi flood water timeseries",
        "source_file": src.name,
        "source_id": src.name,
        "captured": f"{periods[0]['pre_date']} to {periods[-1]['post_date']}",
        "kind": "Estimated",
        "note": (
            "Classed water and flood sample points for pre/post image pairs along the Mithi. "
            "Heatmap is point density, not a surveyed flood outline."
        ),
        "bounds": {
            "west": min(lons),
            "south": min(lats),
            "east": max(lons),
            "north": max(lats),
        },
        "default_period": peak["id"],
        "classes": {
            "water": {"label": "Surface water", "color": "#2f9bd6"},
            "flood": {"label": "Flood water", "color": "#c2372a"},
        },
        "periods": periods,
    }
    (ASSET / "mula-mutha-flood-water.json").write_text(json.dumps(index, indent=2), encoding="utf-8")
    print("flood index default", peak["id"])


def convert_lithology():
    src = DOWNLOADS / "dfe45ce38aae47a7b6a29c9eefd90124.kmz"
    tmp = ROOT / "tmp-lithology-convert"
    if tmp.exists():
        shutil.rmtree(tmp)
    tmp.mkdir()
    with zipfile.ZipFile(src) as zf:
        zf.extractall(tmp)
    kml_path = next(tmp.rglob("*.kml"))
    png_path = next(p for p in tmp.rglob("*.png") if p.name.lower() == "overlay.png")
    legend_path = next((p for p in tmp.rglob("*.png") if p.name.lower() == "legend.png"), None)
    kml = kml_path.read_text(encoding="utf-8", errors="replace")
    box = kml_box(kml)

    # Official KMZ legend order / colours (files/legend.png).
    class_defs = [
        {"id": 0, "key": "655940", "label": "Silty / Sandy Channel Sediment", "color": "#655940"},
        {"id": 1, "key": "8B0000", "label": "Basaltic / Fresh Basalt Spectral Zone", "color": "#8b0000"},
        {"id": 2, "key": "FFD700", "label": "Weathered Basalt", "color": "#ffd700"},
        {"id": 3, "key": "FF0000", "label": "Lateritic / Ferruginous Zone", "color": "#ff0000"},
        {"id": 4, "key": "9370DB", "label": "Clay-Rich / Altered Zone", "color": "#9370db"},
        {"id": 5, "key": "F4A460", "label": "Alluvial / Sandy-Clayey Sediment", "color": "#f4a460"},
        {"id": 6, "key": "00A6A6", "label": "Estuarine / Clayey-Silt Sediment", "color": "#00a6a6"},
        {"id": 7, "key": "808080", "label": "Mixed Weathered Geological Material", "color": "#808080"},
        {"id": 8, "key": "FF69B4", "label": "Exposed / Bright Mineral Surface", "color": "#ff69b4"},
    ]

    out_png = ASSET / "mula-mutha-spectral-lithology.png"
    out_legend = ASSET / "mula-mutha-spectral-lithology-legend.png"
    out_kml = ASSET / "mula-mutha-spectral-lithology.kml"
    shutil.copyfile(png_path, out_png)
    shutil.copyfile(kml_path, out_kml)
    if legend_path:
        shutil.copyfile(legend_path, out_legend)

    # Skip frequency-based relabeling — legend colours are authoritative.
    refined = None
    if False:
        im = Image.open(out_png).convert("RGBA")
        opaque = Counter((r, g, b) for r, g, b, a in im.getdata() if a > 200)
        top = opaque.most_common(20)
        print("lithology top colours:")
        for (r, g, b), n in top[:12]:
            print(f"  #{r:02X}{g:02X}{b:02X}  n={n}")

        refined = []
        used = set()
        labels = [d["label"] for d in class_defs]
        for (r, g, b), n in top:
            if n < 500:
                continue
            key = f"{r:02X}{g:02X}{b:02X}"
            if any(abs(r - ur) + abs(g - ug) + abs(b - ub) < 40 for ur, ug, ub in used):
                continue
            used.add((r, g, b))
            idx = len(refined)
            if idx >= len(labels):
                break
            refined.append(
                {
                    "id": idx,
                    "key": key,
                    "label": labels[idx],
                    "color": f"#{r:02x}{g:02x}{b:02x}",
                }
            )
    if refined and len(refined) >= 5:
        class_defs = refined

    sampled = sample_overlay(out_png, class_defs, max_d=55)
    area = box_area_ha(box)
    for c in sampled["classes"]:
        c["area_ha"] = round(area * (c["share_pct"] / 100.0), 1) if sampled["total_pixels"] else 0.0

    meta = {
        "name": "Mithi River Lithological Interpretation",
        "theme": "Geology",
        "kind": "Estimated",
        "note": (
            "Provisional spectral lithology from Mithi_River_Lithology_Interpretation_Map.tif. "
            "Class 0 rendered as Silty/Sandy Channel Sediment (not open-water blue)."
        ),
        "bounds": box,
        "raster": "/asset/mula-mutha-spectral-lithology.png",
        "legend": "/asset/mula-mutha-spectral-lithology-legend.png",
        "imageCoordinates": coords_from_box(box),
        "classes": sampled["classes"],
        "total_pixels": sampled["total_pixels"],
        "total_area_ha": round(area, 1),
        "width": sampled["width"],
        "height": sampled["height"],
        "source_kmz": src.name,
        "source_name": "Mithi River - Lithological Interpretation",
    }
    (ASSET / "mula-mutha-spectral-lithology.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
    print("lithology classes", len(sampled["classes"]))


def convert_erosion():
    # Already unpacked in _ingest/zip2
    src_dir = INGEST / "zip2"
    kml_path = src_dir / "Erosion_Hotspots_2016_2026.kml"
    png_path = src_dir / "Erosion_Hotspots_2016_2026_overlay.png"
    legend_path = src_dir / "legend.png"
    kml = kml_path.read_text(encoding="utf-8")
    box = kml_box(kml)
    out_png = ASSET / "mula-mutha-erosion-hotspots.png"
    out_legend = ASSET / "mula-mutha-erosion-hotspots-legend.png"
    out_kml = ASSET / "mula-mutha-erosion-hotspots.kml"
    shutil.copyfile(png_path, out_png)
    shutil.copyfile(kml_path, out_kml)
    if legend_path.exists():
        shutil.copyfile(legend_path, out_legend)

    class_defs = [
        {"id": 0, "key": "FFFFFF", "label": "No / very low erosion", "color": "#ffffff"},
        {"id": 1, "key": "FFFF00", "label": "Low erosion", "color": "#ffff00"},
        {"id": 2, "key": "FFA500", "label": "Moderate erosion", "color": "#ffa500"},
        {"id": 3, "key": "FF0000", "label": "High erosion", "color": "#ff0000"},
        {"id": 4, "key": "800000", "label": "Very high erosion", "color": "#800000"},
    ]
    # Also accept light-green no-erosion if present
    im = Image.open(out_png).convert("RGBA")
    opaque = Counter((r, g, b) for r, g, b, a in im.getdata() if a > 0)
    print("erosion top colours:")
    for (r, g, b), n in opaque.most_common(10):
        print(f"  #{r:02X}{g:02X}{b:02X}  n={n}")

    sampled = sample_overlay(out_png, class_defs, max_d=60)
    area = box_area_ha(box)
    for c in sampled["classes"]:
        c["area_ha"] = round(area * (c["share_pct"] / 100.0), 1)

    meta = {
        "name": "Bank Erosion Hotspot (2016–2026)",
        "theme": "Geology",
        "period": "2016-2026",
        "sensor": "Year-to-year erosion count, clipped to Mithi AOI",
        "note": "Hotspot = number of year-to-year periods with detected erosion.",
        "bounds": box,
        "raster": "/asset/mula-mutha-erosion-hotspots.png",
        "legend": "/asset/mula-mutha-erosion-hotspots-legend.png",
        "imageCoordinates": coords_from_box(box),
        "classes": sampled["classes"],
        "total_pixels": sampled["total_pixels"],
        "total_area_ha": round(area, 1),
        "source_zip": "2edcb32fc4504446a4740eb43dd27ede.zip",
        "source_name": "Bank Erosion Hotspot - Mithi River (2016-2026)",
    }
    (ASSET / "mula-mutha-erosion-hotspots.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
    print("erosion done")


def convert_silt():
    src_dir = INGEST / "zip1"
    class_defs = [
        {"id": 1, "key": "44CE1B", "label": "Low", "color": "#44ce1b"},
        {"id": 2, "key": "CEDD54", "label": "Moderate", "color": "#cedd54"},
        {"id": 3, "key": "F3B549", "label": "High", "color": "#f3b549"},
        {"id": 4, "key": "E51F1F", "label": "Very High", "color": "#e51f1f"},
    ]
    month_files = [
        ("2026-01", "Jan 2026", "Mithi_River_2026_01_January_Silt_Classification"),
        ("2026-02", "Feb 2026", "Mithi_River_2026_02_February_Silt_Classification"),
        ("2026-03", "Mar 2026", "Mithi_River_2026_03_March_Silt_Classification"),
        ("2026-04", "Apr 2026", "Mithi_River_2026_04_April_Silt_Classification"),
        ("2026-05", "May 2026", "Mithi_River_2026_05_May_Silt_Classification"),
        ("2026-06", "Jun 2026", "Mithi_River_2026_06_June_Silt_Classification"),
        ("2026-07", "Jul 2026", "Mithi_River_2026_07_July_Silt_Classification"),
    ]
    # clear old volume overlays (Mithi pack has class only)
    for old in ASSET.glob("mula-mutha-silt-volume-*.png"):
        old.unlink()

    periods = []
    for i, (month, label, stem) in enumerate(month_files):
        kml_path = src_dir / f"{stem}.kml"
        png_path = src_dir / f"{stem}_overlay.png"
        kml = kml_path.read_text(encoding="utf-8")
        box = kml_box(kml)
        out_png = ASSET / f"mula-mutha-silt-class-{month}.png"
        shutil.copyfile(png_path, out_png)
        sampled = sample_overlay(out_png, class_defs, max_d=55)
        area = box_area_ha(box)
        classes = []
        for c in sampled["classes"]:
            classes.append(
                {
                    **c,
                    "area_ha": round(area * (c["share_pct"] / 100.0), 1),
                }
            )
        periods.append(
            {
                "id": i,
                "month": month,
                "label": label,
                "bounds": box,
                "imageCoordinates": coords_from_box(box),
                "classification": f"/asset/mula-mutha-silt-class-{month}.png",
                "volume": None,
                "classes": classes,
                "total_pixels": sampled["total_pixels"],
                "classified_area_ha": round(
                    sum(c["area_ha"] for c in classes), 1
                ),
            }
        )
        print("silt", month, "px", sampled["total_pixels"])

    legend = src_dir / "legend.png"
    if legend.exists():
        shutil.copyfile(legend, ASSET / "mula-mutha-silt-legend.png")

    index = {
        "name": "Mithi River silt classification",
        "source_folder": "Mithi silt Jan–Jul 2026",
        "source_id": "f2b2e1eb0ab84942adbede8b8d8f9488.zip",
        "captured": "2026-01 to 2026-07",
        "kind": "Estimated",
        "sensor": "Raster silt classification (monthly GroundOverlays)",
        "note": "Discrete silt classes Low–Very High. No volume surface supplied for Mithi.",
        "default_period": 5,
        "classes": class_defs,
        "csv": None,
        "periods": periods,
    }
    (ASSET / "mula-mutha-silt.json").write_text(json.dumps(index, indent=2), encoding="utf-8")
    print("silt index periods", len(periods))


def convert_salinity():
    """Ship relative salinity as a GroundOverlay raster (same path pattern as silt)."""
    src_dir = INGEST / "kmz2"
    kml_path = src_dir / "doc.kml"
    png_path = src_dir / "Mithi_Relative_Salinity_Classes.png"
    legend_path = src_dir / "Mithi_Relative_Salinity_Classes_legend.png"
    kml = kml_path.read_text(encoding="utf-8")
    box = kml_box(kml)

    out_png = ASSET / "mithi-salinity-overlay.png"
    out_legend = ASSET / "mithi-salinity-legend.png"
    out_kml = ASSET / "mithi-salinity.kml"
    shutil.copyfile(png_path, out_png)
    shutil.copyfile(kml_path, out_kml)
    if legend_path.exists():
        shutil.copyfile(legend_path, out_legend)

    im = Image.open(out_png).convert("RGBA")
    opaque = Counter((r, g, b) for r, g, b, a in im.getdata() if a > 0)
    print("salinity top colours:")
    for (r, g, b), n in opaque.most_common(10):
        print(f"  #{r:02X}{g:02X}{b:02X}  n={n}")

    # Convention 1=Low .. 5=High — refine keys from top colours
    class_defs = [
        {"id": 1, "key": "0000FF", "label": "Low", "color": "#0000ff"},
        {"id": 2, "key": "00BFFF", "label": "Moderately low", "color": "#00bfff"},
        {"id": 3, "key": "00FF00", "label": "Moderate", "color": "#00ff00"},
        {"id": 4, "key": "FFFF00", "label": "Moderately high", "color": "#ffff00"},
        {"id": 5, "key": "FF0000", "label": "High", "color": "#ff0000"},
    ]
    # replace keys with actual top colours ordered by count if 5+ distinct
    tops = []
    used = set()
    for (r, g, b), n in opaque.most_common(30):
        if n < 50:
            continue
        if any(abs(r - ur) + abs(g - ug) + abs(b - ub) < 35 for ur, ug, ub in used):
            continue
        used.add((r, g, b))
        tops.append((r, g, b, n))
        if len(tops) >= 5:
            break
    if len(tops) >= 3:
        # Sort by hue-ish (blue→red) using red-blue difference as salinity proxy
        tops_sorted = sorted(tops, key=lambda t: t[0] - t[2])
        for i, (r, g, b, n) in enumerate(tops_sorted[:5]):
            class_defs[i]["key"] = f"{r:02X}{g:02X}{b:02X}"
            class_defs[i]["color"] = f"#{r:02x}{g:02x}{b:02x}"

    sampled = sample_overlay(out_png, class_defs, max_d=55)
    area = box_area_ha(box)
    for c in sampled["classes"]:
        c["area_ha"] = round(area * (c["share_pct"] / 100.0), 1)

    meta = {
        "name": "Mithi Relative Salinity Classes",
        "kind": "Estimated",
        "note": "Relative salinity classification, water pixels only. Class order Low→High by convention.",
        "bounds": box,
        "raster": "/asset/mithi-salinity-overlay.png",
        "legend": "/asset/mithi-salinity-legend.png",
        "imageCoordinates": coords_from_box(box),
        "classes": sampled["classes"],
        "total_pixels": sampled["total_pixels"],
        "total_area_ha": round(area, 1),
        "source_kmz": "b5b97e2c19f348d29eb9605098729fcd.kmz",
    }
    (ASSET / "mithi-salinity.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

    # Also write a lightweight ndsi-salinity.json so Layers panel can show legend;
    # map wiring for the raster is added in MapComponent.
    ndsi_meta = {
        "name": "Mithi Relative Salinity Classes",
        "kind": "Estimated",
        "mode": "raster",
        "raster": "/asset/mithi-salinity-overlay.png",
        "bounds": box,
        "imageCoordinates": coords_from_box(box),
        "classes": [
            {"class": c["id"], "label": c["label"], "color": c["color"], "share_pct": c["share_pct"]}
            for c in sampled["classes"]
        ],
        "feature_count": 0,
        "geojson": None,
    }
    (ASSET / "mula-mutha-ndsi-salinity.json").write_text(json.dumps(ndsi_meta, indent=2), encoding="utf-8")
    # empty geojson so old polygon layer does not error
    empty = {"type": "FeatureCollection", "features": [], "name": "Mithi salinity (raster)"}
    (ASSET / "mula-mutha-ndsi-salinity.geojson").write_text(json.dumps(empty), encoding="utf-8")
    print("salinity done")


def write_bounds_hint():
    """Default image coords for MapComponent fallback — Mithi reach."""
    box = {
        "north": 19.094260871183593,
        "south": 19.038481978963834,
        "east": 72.88661777120646,
        "west": 72.82722078868055,
    }
    (ASSET / "mithi-bounds.json").write_text(
        json.dumps(
            {
                "name": "Mithi River AOI",
                "bounds": box,
                "imageCoordinates": coords_from_box(box),
                "center": [72.8569, 19.0664],
                "zoom": 12.5,
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def main():
    convert_river()
    convert_buffer()
    convert_garbage()
    counts = convert_tributaries()
    convert_flood_water()
    convert_lithology()
    convert_erosion()
    convert_silt()
    convert_salinity()
    write_bounds_hint()
    # Persist tributary counts for layerLegends update
    (ROOT / "scripts" / "_mithi_tributary_counts.json").write_text(
        json.dumps(dict(counts), indent=2), encoding="utf-8"
    )
    print("INGEST COMPLETE")


if __name__ == "__main__":
    main()
